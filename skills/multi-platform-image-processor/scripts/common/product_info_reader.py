from __future__ import annotations

import logging
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable, Iterator

from openpyxl import load_workbook
import xlrd

from .nas_paths import list_files_fast
from .product_matcher import MatchResult, contains_exact_code, normalize_identity


logger = logging.getLogger(__name__)
EXCEL_SUFFIXES = {".xlsx", ".xls", ".xlsm"}
DEFAULT_ALIASES = {
    "产品货号": {"产品货号", "货号", "款号", "商品货号"},
    "产品名称": {"产品名称", "品名", "商品名称", "名称"},
    "中文面料": {"中文面料", "中文面料信息", "面料", "成分", "材质", "面料成分"},
    "颜色": {"颜色", "颜色名称", "色名"},
    "尺码": {"尺码", "规格", "规格尺码"},
}


@dataclass(frozen=True)
class ProductInfoRecord:
    """保存 Excel 中的一条产品记录。"""

    file: Path
    sheet: str
    row: int
    data: dict[str, Any]

    def get(self, field: str, default: Any = "") -> Any:
        """按规范字段名返回单元格值。"""
        return self.data.get(field, default)


def extract_chinese_material(value: object) -> str:
    """从可能包含双语的单元格中提取中文面料原文。

    参数：
        value：产品信息单元格原始值。
    返回值：
        英文面料段之前的中文行，保持中文内容和换行顺序。
    """
    chinese_lines: list[str] = []
    for raw_line in str(value or "").replace("\r\n", "\n").replace("\r", "\n").split("\n"):
        line = raw_line.strip()
        if not line:
            continue
        if re.search(r"[A-Za-z]", line):
            break
        chinese_lines.append(line)
    return "\n".join(chinese_lines).strip()


def canonical_header(value: object, aliases: dict[str, set[str]]) -> str:
    """将表头映射到规范字段名。"""
    normalized = str(value or "").strip().replace(" ", "")
    for field, names in aliases.items():
        if normalized in {name.replace(" ", "") for name in names}:
            return field
    return normalized


def _records_from_rows(
    file: Path,
    sheet: str,
    rows: Iterable[list[Any]],
    aliases: dict[str, set[str]],
) -> Iterator[ProductInfoRecord]:
    buffered = list(rows)
    header_index = next(
        (
            index for index, row in enumerate(buffered[:50])
            if any(canonical_header(value, aliases) == "产品货号" for value in row)
        ),
        None,
    )
    if header_index is None:
        return
    headers = [canonical_header(value, aliases) for value in buffered[header_index]]
    for index, row in enumerate(buffered[header_index + 1 :], start=header_index + 2):
        data = {
            header: value
            for header, value in zip(headers, row)
            if header and value not in (None, "")
        }
        if data.get("产品货号") not in (None, ""):
            yield ProductInfoRecord(file, sheet, index, data)


def read_product_records(
    path: Path,
    aliases: dict[str, set[str]] | None = None,
) -> list[ProductInfoRecord]:
    """读取 Excel 文件中的产品记录。

    参数：
        path：`.xlsx`、`.xls` 或 `.xlsm` 文件路径。
        aliases：可选的字段表头别名。
    返回值：
        文件中识别到的全部产品记录。
    """
    field_aliases = aliases or DEFAULT_ALIASES
    suffix = path.suffix.lower()
    if suffix not in EXCEL_SUFFIXES:
        raise RuntimeError(f"不支持的产品信息格式：{path}")
    records: list[ProductInfoRecord] = []
    logger.info("开始读取产品信息 file=%r", str(path))
    if suffix in {".xlsx", ".xlsm"}:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            for worksheet in workbook.worksheets:
                rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
                records.extend(_records_from_rows(path, worksheet.title, rows, field_aliases))
        finally:
            workbook.close()
    else:
        workbook = xlrd.open_workbook(path, on_demand=True)
        try:
            for worksheet in workbook.sheets():
                rows = [worksheet.row_values(index) for index in range(worksheet.nrows)]
                records.extend(_records_from_rows(path, worksheet.name, rows, field_aliases))
        finally:
            workbook.release_resources()
    logger.info("产品信息读取完成 file=%r records=%d", str(path), len(records))
    return records


def find_product_info(
    root: Path,
    product_code: str,
    product_name: str = "",
) -> MatchResult:
    """在产品信息目录中查找唯一产品记录。

    参数：
        root：产品信息目录。
        product_code：需要精确匹配的产品货号。
        product_name：用于复核的产品名称。
    返回值：
        唯一产品记录或候选冲突信息。
    """
    matches: list[ProductInfoRecord] = []
    target_code = normalize_identity(product_code)
    target_name = normalize_identity(product_name)
    named_files = [
        path for path in list_files_fast(root, EXCEL_SUFFIXES, product_code)
        if not path.name.startswith("~$")
        if contains_exact_code(str(path.relative_to(root)), product_code)
    ]
    if not named_files and product_name:
        named_files = [
            path for path in list_files_fast(root, EXCEL_SUFFIXES, product_name)
            if not path.name.startswith("~$")
        ]
    logger.info(
        "开始匹配产品信息 code=%s filename_matches=%d",
        product_code,
        len(named_files),
    )
    scan_groups = [named_files]
    for group_index, group in enumerate(scan_groups):
        for path in group:
            try:
                records = read_product_records(path)
            except Exception as exc:
                logger.warning("产品信息文件读取失败 file=%r error=%s", str(path), exc)
                continue
            for record in records:
                if normalize_identity(record.get("产品货号")) != target_code:
                    continue
                if target_name:
                    record_name = normalize_identity(record.get("产品名称"))
                    if record_name and target_name != record_name:
                        continue
                matches.append(record)
        if matches:
            break
        if group_index == 0:
            fallback = [
                path for path in list_files_fast(root, EXCEL_SUFFIXES)
                if not path.name.startswith("~$") and path not in named_files
            ]
            scan_groups.append(fallback)
    if len(matches) == 1:
        return MatchResult(matches[0], list(matches), "产品信息记录唯一")
    if not matches:
        return MatchResult(None, [], "没有找到产品货号精确匹配记录")
    return MatchResult(None, list(matches), "同一产品匹配到多个 Excel 记录")
