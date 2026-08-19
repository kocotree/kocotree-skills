from __future__ import annotations

import json
import os
import unittest
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest.mock import patch

import xlwt
from openpyxl import Workbook
from PIL import Image

from common.color_naming import resolve_color_names
from common.font_assets import load_font_assets, require_glyphs
from common.nas_paths import require_accessible_directory, to_unc_path
from common.product_info_reader import (
    extract_chinese_material,
    extract_representative_color,
    find_product_info,
    read_product_records,
)
from common.product_matcher import (
    contains_exact_code,
    infer_product_code,
    select_bartender_file,
)
from common.settings import resolve_business_paths
from common.utils import build_platform_directory_names, 平台模板目录名


class BusinessSettingsTests(unittest.TestCase):
    """验证业务路径配置和 UNC 路径归一化。"""

    def test_environment_path_has_highest_priority(self) -> None:
        """验证环境变量中的 NAS 路径覆盖配置文件。"""
        with TemporaryDirectory() as temp_dir_value:
            config = Path(temp_dir_value) / "paths.json"
            config.write_text(
                json.dumps(
                    {
                        "NAS根目录": r"\\config\share",
                        "产品信息相对目录": r"产品中心\产品信息",
                        "合格证相对目录": r"产品中心\合格证",
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )
            with patch.dict(os.environ, {"KOCOTREE_NAS_ROOT": r"\\env\share"}):
                paths = resolve_business_paths(config_path=config)

            self.assertEqual(str(paths.nas_root).rstrip("\\"), r"\\env\share")
            self.assertIn("产品信息", str(paths.product_info_root))

    def test_drive_path_is_converted_to_unc(self) -> None:
        """验证映射盘路径转换为同一共享目录的 UNC 路径。"""
        converted = to_unc_path(
            r"Z:\产品中心\产品信息",
            {"Z:": r"\\192.168.110.20\浙江酷趣"},
        )

        self.assertEqual(
            str(converted),
            r"\\192.168.110.20\浙江酷趣\产品中心\产品信息",
        )

    def test_inaccessible_directory_reports_label(self) -> None:
        """验证不可访问目录返回带业务名称的错误。"""
        with TemporaryDirectory() as temp_dir_value:
            missing = Path(temp_dir_value) / "不存在"
            with self.assertRaisesRegex(RuntimeError, "产品信息目录不可访问"):
                require_accessible_directory(missing, "产品信息目录")

    def test_jd_directory_name_uses_excel_identity(self) -> None:
        """验证京东目录名包含 Excel 货号、空格和产品名称。"""
        names = build_platform_directory_names("P100", "儿童外套")

        self.assertEqual(names["jd"], "P100 儿童外套-京东")

    def test_jd_template_uses_product_identity_placeholder(self) -> None:
        """验证京东静态模板使用产品身份占位目录名。"""
        self.assertEqual(平台模板目录名["jd"], "产品货号 产品名称-京东")
        template_root = Path(__file__).resolve().parents[2] / "template"
        self.assertTrue((template_root / 平台模板目录名["jd"]).is_dir())
        self.assertFalse((template_root / "京东").exists())

    def test_white_and_transparent_images_use_sku_color_names(self) -> None:
        """验证数字素材通过视觉映射使用 SKU 颜色名称。"""
        with TemporaryDirectory() as temp_dir_value:
            root = Path(temp_dir_value)
            for relative in (
                "SKU/800/豆蔻紫.jpg",
                "SKU/800/远海蓝.jpg",
                "白底图/1.jpg",
                "白底图/2.jpg",
                "透明图/1.png",
                "透明图/2.png",
            ):
                path = root / relative
                path.parent.mkdir(parents=True, exist_ok=True)
                Image.new("RGBA" if path.suffix == ".png" else "RGB", (8, 8), "white").save(path)

            names = resolve_color_names(
                root,
                {
                    "白底图/1.jpg": "豆蔻紫",
                    "白底图/2.jpg": "远海蓝",
                    "透明图/1.png": "豆蔻紫",
                    "透明图/2.png": "远海蓝",
                },
            )

            self.assertEqual(names[(root / "白底图/1.jpg").resolve()], "豆蔻紫")
            self.assertEqual(names[(root / "透明图/2.png").resolve()], "远海蓝")

    def test_gift_sku_uses_only_no_gift_names_for_color_assets(self) -> None:
        """验证赠品结构的颜色素材只接受无赠品规格名称。"""
        with TemporaryDirectory() as temp_dir_value:
            root = Path(temp_dir_value)
            for relative in (
                "SKU/加赠品/800/赠品款.jpg",
                "SKU/无赠品/800/企鹅团团-浅灰.jpg",
                "白底图/1.jpg",
            ):
                path = root / relative
                path.parent.mkdir(parents=True, exist_ok=True)
                Image.new("RGB", (8, 8), "white").save(path)

            names = resolve_color_names(
                root,
                {"白底图/1.jpg": "企鹅团团-浅灰"},
            )

            self.assertEqual(
                names[(root / "白底图/1.jpg").resolve()],
                "企鹅团团-浅灰",
            )

    def test_offsite_template_uses_business_folder_names(self) -> None:
        """验证站外模板只包含最终业务目录名称。"""
        template_root = Path(__file__).resolve().parents[2] / "template" / "站外通用版"
        actual = {path.name for path in template_root.iterdir() if path.is_dir()}

        self.assertEqual(
            actual,
            {"详情页", "sku", "白底图", "白底图＋logo", "透明图", "主图", "素材图"},
        )


class ProductInfoTests(unittest.TestCase):
    """验证三种 Excel 格式读取和产品唯一匹配。"""

    @staticmethod
    def _write_openxml(path: Path) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "产品资料"
        sheet.append(["说明"])
        sheet.append(["产品货号", "产品名称", "中文面料信息", "颜色", "尺码"])
        sheet.append(["KQ26143", "儿童长裤", "棉 95% 氨纶 5%", "蓝色", "110"])
        workbook.save(path)

    @staticmethod
    def _write_xls(path: Path) -> None:
        workbook = xlwt.Workbook()
        sheet = workbook.add_sheet("产品资料")
        values = [
            ["产品货号", "产品名称", "中文面料", "颜色", "尺码"],
            ["KQ25143", "儿童外套", "聚酯纤维 100%", "红色", "110"],
        ]
        for row_index, row in enumerate(values):
            for column_index, value in enumerate(row):
                sheet.write(row_index, column_index, value)
        workbook.save(str(path))

    @staticmethod
    def _write_specification_xls(path: Path) -> None:
        workbook = xlwt.Workbook()
        sheet = workbook.add_sheet("产品资料")
        values = [
            ["货号", "品名", "规格", "尺码", "成分"],
            ["KQ26169", "小云暖抓绒马甲（按扣款）", "星夜蓝100", "100/48", "100%聚酯纤维"],
        ]
        for row_index, row in enumerate(values):
            for column_index, value in enumerate(row):
                sheet.write(row_index, column_index, value)
        workbook.save(str(path))

    def test_reads_xlsx_xlsm_and_xls(self) -> None:
        """验证三种产品信息文件均可读取中文面料。"""
        with TemporaryDirectory() as temp_dir_value:
            root = Path(temp_dir_value)
            xlsx = root / "KQ26143.xlsx"
            xlsm = root / "KQ26143.xlsm"
            xls = root / "KQ25143.xls"
            self._write_openxml(xlsx)
            self._write_openxml(xlsm)
            self._write_xls(xls)

            for path in (xlsx, xlsm, xls):
                records = read_product_records(path)
                self.assertEqual(len(records), 1)
                self.assertTrue(records[0].get("中文面料"))

    def test_multiple_excel_records_block_automatic_selection(self) -> None:
        """验证同货号存在多份产品信息时不自动猜测。"""
        with TemporaryDirectory() as temp_dir_value:
            root = Path(temp_dir_value)
            self._write_openxml(root / "KQ26143-A.xlsx")
            self._write_openxml(root / "KQ26143-B.xlsx")

            result = find_product_info(root, "KQ26143", "儿童长裤")

            self.assertIsNone(result.selected)
            self.assertEqual(len(result.candidates), 2)
            self.assertIn("多个", result.reason)

    def test_bilingual_cell_returns_only_chinese_material(self) -> None:
        """验证中英文同单元格时只提取英文段之前的中文原文。"""
        value = "成分：98.3%聚酯纤维\n1.7%氨纶\ncomponent:98.3%polyester\n1.7%spandex"

        self.assertEqual(extract_chinese_material(value), "成分：98.3%聚酯纤维\n1.7%氨纶")

    def test_chinese_material_keeps_embedded_latin_symbols(self) -> None:
        """验证中文面料中的化学符号不会被误判为英文段。"""
        value = "杯身内胆：06Cr17Ni12Mo2(内胆S316)\n螺牙盖/吸管：聚丙烯（PP）\nMaterial: stainless steel"

        self.assertEqual(
            extract_chinese_material(value),
            "杯身内胆：06Cr17Ni12Mo2(内胆S316)\n螺牙盖/吸管：聚丙烯（PP）",
        )

    def test_specification_and_size_remain_separate(self) -> None:
        """验证真实产品表中的规格与尺码分别保留。"""
        with TemporaryDirectory() as temp_dir_value:
            path = Path(temp_dir_value) / "KQ26169.xls"
            self._write_specification_xls(path)

            record = read_product_records(path)[0]

        self.assertEqual(record.get("规格"), "星夜蓝100")
        self.assertEqual(record.get("尺码"), "100/48")
        self.assertEqual(extract_representative_color(record), "星夜蓝")


class ProductMatcherTests(unittest.TestCase):
    """验证货号识别和 BarTender 产品目录、颜色、尺码选择。"""

    def test_product_code_requires_exact_boundary(self) -> None:
        """验证相邻的更长货号不会被当作精确匹配。"""
        self.assertTrue(contains_exact_code("KQ26143 产品", "KQ26143"))
        self.assertFalse(contains_exact_code("KQ261430 产品", "KQ26143"))

    def test_product_code_can_be_inferred_from_product_directory(self) -> None:
        """验证产品目录中唯一货号可供完整流程识别。"""
        self.assertEqual(infer_product_code(Path("KQ26143 儿童长裤") / "数据包"), "KQ26143")

    def test_prefers_110_then_smallest_size(self) -> None:
        """验证优先 110 码且缺少时选择最小尺码。"""
        with TemporaryDirectory() as temp_dir_value:
            root = Path(temp_dir_value)
            product = root / "儿童长裤"
            product.mkdir()
            for name in ("儿童长裤 蓝色100.btw", "儿童长裤 蓝色110.btw"):
                (product / name).write_bytes(b"btw")
            preferred = select_bartender_file(root, "儿童长裤", "蓝色")
            (product / "儿童长裤 蓝色110.btw").unlink()
            fallback = select_bartender_file(root, "儿童长裤", "蓝色")

        self.assertEqual(preferred.selected.name, "儿童长裤 蓝色110.btw")
        self.assertEqual(fallback.selected.name, "儿童长裤 蓝色100.btw")
        self.assertIn("最小尺码 100", fallback.reason)

    def test_only_top_level_bartender_file_is_selected(self) -> None:
        """验证默认匹配不进入特殊版本子目录。"""
        with TemporaryDirectory() as temp_dir_value:
            root = Path(temp_dir_value)
            product = root / "儿童长裤"
            special = product / "委托商版"
            special.mkdir(parents=True)
            (product / "儿童长裤 蓝色110.btw").write_bytes(b"standard")
            (special / "儿童长裤 蓝色110.btw").write_bytes(b"special")

            result = select_bartender_file(root, "儿童长裤", "蓝色")

        self.assertEqual(result.selected, product / "儿童长裤 蓝色110.btw")
        self.assertEqual(len(result.candidates), 1)

    def test_unique_alpha_size_bartender_file_is_selected(self) -> None:
        """验证唯一颜色候选使用字母尺码时仍可安全选择。"""
        with TemporaryDirectory() as temp_dir_value:
            root = Path(temp_dir_value)
            product = root / "儿童帽子"
            product.mkdir()
            expected = product / "儿童帽子 深灰XS.btw"
            expected.write_bytes(b"btw")

            result = select_bartender_file(root, "儿童帽子", "深灰XS")

        self.assertEqual(result.selected, expected)
        self.assertIn("唯一文件", result.reason)

    def test_missing_product_name_directory_is_blocked(self) -> None:
        """验证正式产品名称没有对应一级目录时停止选择。"""
        with TemporaryDirectory() as temp_dir_value:
            root = Path(temp_dir_value)
            (root / "儿童长裤").mkdir()

            result = select_bartender_file(root, "儿童外套", "蓝色")

        self.assertIsNone(result.selected)
        self.assertIn("合格证目录", result.reason)


class FontAssetTests(unittest.TestCase):
    """验证 Skill 字体资产。"""

    def test_all_font_assets_load_and_cover_roles(self) -> None:
        """验证四个字体文件的内部名称和必需字形。"""
        assets = load_font_assets()

        self.assertEqual(set(assets), {"方正兰亭中黑", "方正兰亭中黑备用", "G8321", "数字1"})
        require_glyphs(assets["方正兰亭中黑"], "面料：棉95%")
        require_glyphs(assets["G8321"], "023456789.")
        require_glyphs(assets["数字1"], "1")
        with self.assertRaisesRegex(RuntimeError, "缺少字形"):
            require_glyphs(assets["G8321"], "%")


if __name__ == "__main__":
    unittest.main()
